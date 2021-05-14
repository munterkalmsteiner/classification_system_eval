import glob
import xlrd
import re
import csv
import json
from openpyxl import load_workbook
import unicodedata as ud
from treelib import Node, Tree

ROOT_NAME = "root"
BASE_PATH = "classification_systems"

def uniclass(name):
    path = f'{BASE_PATH}/uniclass'

    tree = Tree()
    tree.create_node(name, ROOT_NAME)
    files = glob.glob(f'{path}/Uniclass2015*.xlsx')

    for file in files:
        wb = load_workbook(file)
        ws = wb.active
        data = ws.values
        table_name = next(data)[0].strip()
        table_id = table_name[:2].strip()
        tree.create_node(table_name, table_id, ROOT_NAME)

        #skip to the data rows
        next(data)
        next(data)

        for row in data:
            identifier = row[0].strip()
            content = row[5].strip()
            parent = identifier[:-3]
            tree.create_node(content, identifier, parent)

    return tree

def omniclass_drop_trailing_zeros(o_number):
    components = re.split(' ', o_number)
    no_zeros = [comp for comp in components if comp != "00"]
    return ' '.join(no_zeros)

def get_omniclass_parent(o_number):
    components = re.split(' |-', o_number)
    parent = components[:-1]

    if len(parent) == 1:
        return parent[0]
    if len(parent) == 2:
        return '-'.join(parent[:2])
    else:
        return '-'.join(parent[:2]) + ' ' + ' '.join(parent[2:])

def omniclass(name):
    # Tables 14 (Spaces by form) and 35 (Tools) are only available as pdf, hence skipped for now.

    # OmniClass FLAT had the following typos which I fixed manually by checking with the non-FLAT sheet
    # - Table 23 (11 removed at the end)
    #   23-17 19 13 25 13 11 -> 23-17 19 13 25 13
    #   23-17 19 13 25 15 11 -> 23-17 19 13 25 15
    #   23-17 19 13 25 17 11 -> 23-17 19 13 25 17
    #
    # - Table 23 (duplicates)
    #   23-39 11 15 17 -> 23-39 11 15 21
    #   23-39 11 15 19 -> 23-39 11 15 23
    #   23-39 11 15 21 -> 23-39 11 15 25
    #
    # - Table 23 (1 missing)
    #   23-39 29 13 2 19 -> 23-39 29 13 21 19
    #
    # - Table 23 (9 removed)
    #   23-39 29 13 21 9 23 -> 23-39 29 13 21 23
    #
    # - Table 36 (replaced 31 with 26)
    #   36-71 31 21 19 13 -> 36-71 26 21 19 13

    path = f'{BASE_PATH}/omniclass'

    tree = Tree()
    tree.create_node(name, ROOT_NAME)
    files = glob.glob(f'{path}/OmniClass*.xls')

    for file in files:
        #Skip the irregular Phases table which contains anyway only 11 items
        if "OmniClass_31_2012-10-30.xls" in file:
            continue
        #Skip old version of table 22
        if "OmniClass_22_2012-05-16.xls" in file:
            continue

        wb = xlrd.open_workbook(file)
        ws_name = next(sn for sn in wb.sheet_names() if "flat" in sn.lower())
        ws = wb.sheet_by_name(ws_name)
        table_name = ' '.join(ws.row_values(0))
        table_id = ws.cell_value(2, 0)[:2]
        tree.create_node(table_name, table_id, ROOT_NAME)

        for r in range(2, ws.nrows):
            if ws.cell_value(r, 0) == "End of Table":
                continue

            identifier = ws.cell_value(r, 0).strip()
            identifier = ud.normalize("NFKD", identifier)
            identifier = omniclass_drop_trailing_zeros(identifier)

            content = ws.cell_value(r, 1).strip()
            content = ud.normalize("NFKD", content)

            parent = get_omniclass_parent(identifier)
            tree.create_node(content, identifier, parent)

    return tree


def path_to_root(tree, node, path):
    path.append(node)
    if not node.is_root():
        path_to_root(tree, tree.parent(node.identifier), path)
    return path

def naics(name):
    tree = Tree()
    tree.create_node(name, ROOT_NAME)

    #Since there are no tables in this classification, we create an artificial one so the analysis works.
    DUMMY_TABLE = "dummy_table"
    tree.create_node(DUMMY_TABLE, DUMMY_TABLE, ROOT_NAME)
    wb = load_workbook(f'{BASE_PATH}/NAICS/2-6 digit_2017_Codes.xlsx')
    ws = wb.active
    data = ws.values

    #Skipping first two rows
    next(data)
    next(data)

    for row in data:
        identifier = str(row[1])

        #Skip identifiers ending with "0" since these have the same content as their parent.
        #This would introduce artificial leaf nodes. 
        if identifier[-1] == "0":
            continue

        content = row[2].strip()

        #Several codes have been merged, which breaks the regularity of the code construction. 
        if identifier == "31-33":
            tree.create_node(content, "31", DUMMY_TABLE)
            tree.create_node(content, "32", DUMMY_TABLE)
            tree.create_node(content, "33", DUMMY_TABLE)
            continue;

        if identifier == "44-45":
            tree.create_node(content, "44", DUMMY_TABLE)
            tree.create_node(content, "45", DUMMY_TABLE)
            continue;

        if identifier == "48-49":
            tree.create_node(content, "48", DUMMY_TABLE)
            tree.create_node(content, "49", DUMMY_TABLE)
            continue;

        if len(identifier) == 2:
            parent = DUMMY_TABLE
        else:
            parent = identifier[:-1]

        tree.create_node(content, identifier, parent)

    return tree

def nace(name):
    tree = Tree();
    tree.create_node(name, ROOT_NAME)

    #Since there are no tables in this classification, we create an artificial one so the analysis works.
    DUMMY_TABLE = "dummy_table"
    tree.create_node(DUMMY_TABLE, DUMMY_TABLE, ROOT_NAME)

    with open(f'{BASE_PATH}/NACE/NACE_REV2_20210505_094912.csv') as csvfile:
        reader = csv.reader(csvfile, delimiter=',')

        #Skip header
        next(reader)

        for row in reader:
            level = row[1].strip()
            identifier = row[2].strip()
            parent = row[3].strip()
            content = row[4].strip()

            if len(parent) == 0:
                tree.create_node(content, identifier, DUMMY_TABLE)
            else:
                parent_node = tree.get_node(parent)
                #Skip nodes whose parents have the same content and are at the lowest level.
                #This would introduce artificial leaf nodes.
                if parent_node.tag == content and level == "4":
                    continue
                else:
                    tree.create_node(content, identifier, parent)

    return tree

def eucyber(name):
    tree = Tree();
    tree.create_node(name, ROOT_NAME)

    with open(f'{BASE_PATH}/european_cybersecurity_taxonomy/european_cybersecurity_taxonomy_manual_extraction.csv') as csvfile:
        reader = csv.reader(csvfile, delimiter=',')

        #Skip header
        next(reader)

        for row in reader:
            identifier = row[0].strip()
            parent = row[1].strip()
            content = row[2].strip()

            if len(parent) == 0:
                tree.create_node(content, identifier, ROOT_NAME)
            else:
                tree.create_node(content, identifier, parent)

    return tree

def traverse_dict(tree, d, parent=None):
    identifier = d["id"]
    content = d["name"]

    parent = ROOT_NAME if parent is None else parent
    tree.create_node(content, identifier, parent)

    if "children" in d:
        for child in d["children"]:
            traverse_dict(tree, child, identifier)

def mahaini(name):
    tree = Tree()
    tree.create_node(name, ROOT_NAME)

    with open(f'{BASE_PATH}/mahaini_cybersecurity/Taxonomy.json') as jsonfile:
        data = json.load(jsonfile)

    traverse_dict(tree, data)

    return tree


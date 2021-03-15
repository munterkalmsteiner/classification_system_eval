import glob
import xlrd
import re
import csv
import unicodedata as ud
import pandas as pd
from treelib import Node, Tree
from openpyxl import load_workbook
from math import log
from dataclasses import dataclass

ROOT_NAME = "root"
BASE_PATH = "/home/mun/phd/projects/companies/TRV/DCAT/WP2/part2_evaluating_classification_systems/classification_systems"

eval_result = []

@dataclass
class EvalResult:
    classification_system_name: str
    table_name: str
    number_of_categories: int
    number_of_characteristics: int
    conciseness: float
    robustness: float


def get_uniclass_tree():
    path = f'{BASE_PATH}/uniclass'

    tree = Tree()
    tree.create_node("UniClass", ROOT_NAME)
    files = glob.glob(f'{path}/Uniclass2015*.xlsx')

    for file in files:
        wb = load_workbook(file)
        ws = wb.active
        data = ws.values
        table_name = next(data)[0].strip()
        table_id = table_name[:2].strip()
        tree.create_node(table_name, table_id, ROOT_NAME)

        #skip to the data rows
        next(data)
        next(data)

        for row in data:
            identifier = row[0].strip()
            content = row[5].strip()
            parent = identifier[:-3]
            tree.create_node(content, identifier, parent)

    return tree

def omniclass_drop_trailing_zeros(o_number):
    components = re.split(' ', o_number)
    no_zeros = [comp for comp in components if comp != "00"]
    return ' '.join(no_zeros)

def get_omniclass_parent(o_number):
    components = re.split(' |-', o_number)
    parent = components[:-1]

    if len(parent) == 1:
        return parent[0]
    if len(parent) == 2:
        return '-'.join(parent[:2])
    else:
        return '-'.join(parent[:2]) + ' ' + ' '.join(parent[2:])

def get_omniclass_tree():
    # Tables 14 (Spaces by form) and 35 (Tools) are only available as pdf, hence skipped for now.

    # OmniClass FLAT had the following typos which I fixed manually by checking with the non-FLAT sheet
    # - Table 23 (11 removed at the end)
    #   23-17 19 13 25 13 11 -> 23-17 19 13 25 13
    #   23-17 19 13 25 15 11 -> 23-17 19 13 25 15
    #   23-17 19 13 25 17 11 -> 23-17 19 13 25 17
    #
    # - Table 23 (duplicates)
    #   23-39 11 15 17 -> 23-39 11 15 21
    #   23-39 11 15 19 -> 23-39 11 15 23
    #   23-39 11 15 21 -> 23-39 11 15 25
    #
    # - Table 23 (1 missing)
    #   23-39 29 13 2 19 -> 23-39 29 13 21 19
    #
    # - Table 23 (9 removed)
    #   23-39 29 13 21 9 23 -> 23-39 29 13 21 23
    #
    # - Table 36 (replaced 31 with 26)
    #   36-71 31 21 19 13 -> 36-71 26 21 19 13

    path = f'{BASE_PATH}/omniclass'

    tree = Tree()
    tree.create_node("OmniClass", ROOT_NAME)
    files = glob.glob(f'{path}/OmniClass*.xls')

    for file in files:
        #Skip the irregular Phases table which contains anyway only 11 items
        if "OmniClass_31_2012-10-30.xls" in file:
            continue
        #Skip old version of table 22
        if "OmniClass_22_2012-05-16.xls" in file:
            continue

        wb = xlrd.open_workbook(file)
        ws_name = next(sn for sn in wb.sheet_names() if "flat" in sn.lower())
        ws = wb.sheet_by_name(ws_name)
        table_name = ' '.join(ws.row_values(0))
        table_id = ws.cell_value(2, 0)[:2]
        tree.create_node(table_name, table_id, ROOT_NAME)

        for r in range(2, ws.nrows):
            if ws.cell_value(r, 0) == "End of Table":
                continue

            identifier = ws.cell_value(r, 0).strip()
            identifier = ud.normalize("NFKD", identifier)
            identifier = omniclass_drop_trailing_zeros(identifier)

            content = ws.cell_value(r, 1).strip()
            content = ud.normalize("NFKD", content)

            parent = get_omniclass_parent(identifier)
            tree.create_node(content, identifier, parent)

    return tree

def get_coclass_tree():
    tree = Tree();
    tree.create_node("CoClass", ROOT_NAME)

    # Using the cvs file since I have already added there dummy nodes that are missing
    # in the original for a proper tree:
    # Tillgangsystem: Dummy*
    # Added also dummy IDs: 1, 2, 3
    with open(f'{BASE_PATH}/coclass/coclass_v2019_20200109.csv') as csvfile:
        reader = csv.reader(csvfile, delimiter=';')

        #Skip header
        next(reader)

        for row in reader:
            table_name = row[0].strip()
            identifier = row[1].strip()
            content = row[2].strip()

            if len(identifier) == 0:
                tree.create_node(table_name, table_name, ROOT_NAME)
            elif len(identifier) == 1:
                tree.create_node(content, table_name + identifier, table_name)
            else:
                tree.create_node(content, table_name + identifier, table_name + identifier[:-1])

    return tree

def get_sb11_tree():
    # Needed to modify the original file since three tables were stored in a single sheet and the items
    # of the second and third table were not ordered.

    # Fixed typos:
    # alternativtabell: U261 -> U-261

    path = f'{BASE_PATH}/sb11'

    tree = Tree()
    tree.create_node("SB11", ROOT_NAME)

    wb = load_workbook(f'{path}/SB11 CAD-Lager_ Elementkod_2020-11-27 13_46_19.xlsx')
    for ws_name in wb.sheetnames:
        if ws_name == "Original":
            continue

        tree.create_node(ws_name, ws_name, ROOT_NAME)
        ws = wb[ws_name]
        data = ws.values

        #skip header
        next(data)

        for row in data:
            identifier = str(row[0])
            content = row[1].strip()
            table_name = ws_name

            if table_name == "Byggdelar":
                if len(identifier) == 1:
                    parent = table_name
                else:
                    parent = identifier[:-1]
                tree.create_node(content, identifier, parent)
            elif table_name == "Alternativtabell":
                if identifier[0] == "U": # Parts of the table are hierarchical (U-xxx), the rests isn't
                    identifier = identifier.replace("-", "")

                    if identifier == "U":
                        tree.create_node(content, identifier, table_name)
                    else:
                        tree.create_node(content, identifier, identifier[:-1])
                else: #The non-hierarchical items
                    tree.create_node(content, identifier, table_name)
            elif table_name == "Landskapsinformation":
                #These are all flat items, no hierarchy
                tree.create_node(content, identifier, table_name)
    return tree

def eval_conciseness(tree, level, cs_name = None):
    depth_categories = 0
    depth_characteristics = 0
    number_categories = 0
    number_characteristics = 0

    root_node = tree.root

    for node in tree.all_nodes():
        if node.identifier == root_node:
            name = node.tag
            continue
        depth = tree.depth(node)
        if node.is_leaf():
            depth_characteristics += 1/depth
            number_characteristics += 1
        else:
            depth_categories += 1/depth
            number_categories += 1

    conciseness = 1/(1 + log((depth_categories + depth_characteristics) - 1))

    if level == 0:
        print('Classification system: ', end='')
    elif level == 1:
        print('\tTable: ', end='')

    print(f'{name} | Number of categories/characteristics: {number_categories}/{number_characteristics} | Conciseness: {conciseness}')

    if level == 0:
        #eval_result.update({"classification system": cs_name, "table": "N/A",
        #                    "number of categories": number_categories,
        #                    "number of characteristics": number_characteristics,
        #                    "conciseness": conciseness})
        for table in tree.children(ROOT_NAME):
            subtree = tree.subtree(table.identifier)
            eval_conciseness(subtree, level + 1, name)
    #else:
        #eval_result.update({"classification system": cs_name, "table": name,
        #                    "number of categories": number_categories,
        #                    "number of characteristics": number_characteristics,
        #                    "conciseness": conciseness})

tree_uniclass = get_uniclass_tree()
tree_omniclass = get_omniclass_tree()
tree_coclass = get_coclass_tree()
tree_sb11 = get_sb11_tree()

eval_conciseness(tree_uniclass, 0)
eval_conciseness(tree_omniclass, 0)
eval_conciseness(tree_coclass, 0)
eval_conciseness(tree_sb11, 0)

